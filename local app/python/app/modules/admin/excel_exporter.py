"""
Community Admin - Excel/CSV Export
Xuất dữ liệu community và event thành Excel/CSV
"""
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import csv
from app.utils.crypto import Logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExporter:
    """Export community and event data to Excel/CSV"""
    
    def __init__(self, output_path: str = "./exports"):
        self.output_path = output_path
        self._ensure_output_directory()
    
    def _ensure_output_directory(self):
        """Ensure output directory exists"""
        os.makedirs(self.output_path, exist_ok=True)
    
    def export_communities_excel(
        self,
        communities: List[Dict],
        format_type: str = "xlsx"
    ) -> str:
        """Export all communities to Excel/CSV"""
        Logger.info(f"========== Exporting Communities to {format_type.upper()} ==========", "EXPORT")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format_type == "xlsx" and HAS_OPENPYXL:
            return self._export_communities_xlsx(communities, timestamp)
        else:
            return self._export_communities_csv(communities, timestamp)
    
    def _export_communities_xlsx(
        self,
        communities: List[Dict],
        timestamp: str
    ) -> str:
        """Export communities using openpyxl"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Communities"
        
        # Headers
        headers = ["Community ID", "Name", "Description", "Created Date", "Active Members", "Total Events", "Status"]
        ws.append(headers)
        
        # Format header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for community in communities:
            ws.append([
                community.get('community_id', ''),
                community.get('name', ''),
                community.get('description', ''),
                community.get('created_date', ''),
                community.get('active_members', 0),
                community.get('total_events', 0),
                community.get('status', '')
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        file_path = os.path.join(self.output_path, f"Communities_Master_{timestamp}.xlsx")
        wb.save(file_path)
        
        Logger.success(f"Master community file exported: {file_path}")
        return file_path
    
    def _export_communities_csv(
        self,
        communities: List[Dict],
        timestamp: str
    ) -> str:
        """Export communities as CSV"""
        file_path = os.path.join(self.output_path, f"Communities_Master_{timestamp}.csv")
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(
                f,
                fieldnames=['community_id', 'name', 'description', 'created_date', 
                           'active_members', 'total_events', 'status']
            )
            writer.writeheader()
            writer.writerows(communities)
        
        Logger.success(f"Master community file exported (CSV): {file_path}")
        return file_path
    
    def export_community_detail_excel(
        self,
        community_id: str,
        community_name: str,
        events: List[Dict],
        members: List[Dict] = None,
        format_type: str = "xlsx"
    ) -> str:
        """Export community details with events and members"""
        Logger.info(f"========== Exporting {community_name} Details to {format_type.upper()} ==========", "EXPORT")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = "".join(c if c.isalnum() or c in '_-' else '_' for c in community_name)[:30]
        
        if format_type == "xlsx" and HAS_OPENPYXL:
            return self._export_detail_xlsx(safe_name, events, members, timestamp)
        else:
            return self._export_detail_csv(safe_name, events, members, timestamp)
    
    def _export_detail_xlsx(
        self,
        safe_name: str,
        events: List[Dict],
        members: List[Dict],
        timestamp: str
    ) -> str:
        """Export detail as xlsx"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Events"
        
        # Events sheet
        event_headers = ["Event ID", "Event Name", "Date", "Location", "Status", "Attendees", "Description"]
        ws.append(event_headers)
        
        for header_cell in ws[1]:
            header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_cell.font = Font(bold=True, color="FFFFFF")
        
        for event in events:
            ws.append([
                event.get('event_id', ''),
                event.get('event_name', ''),
                event.get('event_date', ''),
                event.get('location', ''),
                event.get('status', ''),
                event.get('attendees', 0),
                event.get('description', '')
            ])
        
        # Members sheet
        if members:
            ws_members = wb.create_sheet("Members")
            member_headers = ["Member ID", "Name", "Email", "Joined Date", "Status", "Role"]
            ws_members.append(member_headers)
            
            for header_cell in ws_members[1]:
                header_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                header_cell.font = Font(bold=True, color="FFFFFF")
            
            for member in members:
                ws_members.append([
                    member.get('member_id', ''),
                    member.get('name', ''),
                    member.get('email', ''),
                    member.get('joined_date', ''),
                    member.get('status', ''),
                    member.get('role', '')
                ])
        
        file_path = os.path.join(self.output_path, f"{safe_name}_Detail_{timestamp}.xlsx")
        wb.save(file_path)
        
        Logger.success(f"Community detail file exported: {file_path}")
        return file_path
    
    def _export_detail_csv(
        self,
        safe_name: str,
        events: List[Dict],
        members: List[Dict],
        timestamp: str
    ) -> str:
        """Export detail as CSV"""
        events_file = os.path.join(self.output_path, f"{safe_name}_Events_{timestamp}.csv")
        members_file = os.path.join(self.output_path, f"{safe_name}_Members_{timestamp}.csv")
        
        # Export events
        with open(events_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(
                f,
                fieldnames=['event_id', 'event_name', 'event_date', 'location', 'status', 'attendees', 'description']
            )
            writer.writeheader()
            writer.writerows(events)
        
        # Export members
        if members:
            with open(members_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=['member_id', 'name', 'email', 'joined_date', 'status', 'role']
                )
                writer.writeheader()
                writer.writerows(members)
        
        Logger.success(f"Community detail files exported (CSV): {events_file}, {members_file}")
        return events_file
